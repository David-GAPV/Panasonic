import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Colors ---
HEADER_FILL = PatternFill(start_color="2D4285", end_color="2D4285", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
DONE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
NOT_DONE_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
DONE_FONT = Font(name="Arial", size=11, color="2E7D32")
PARTIAL_FONT = Font(name="Arial", size=11, color="F57C00")
NOT_DONE_FONT = Font(name="Arial", size=11, color="D32F2F")
NORMAL_FONT = Font(name="Arial", size=11)
BOLD_FONT = Font(name="Arial", size=11, bold=True)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="2D4285")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

def status_style(status):
    if status == "✅ Done":
        return DONE_FILL, DONE_FONT
    elif status == "⚠️ Partial":
        return PARTIAL_FILL, PARTIAL_FONT
    else:
        return NOT_DONE_FILL, NOT_DONE_FONT

# ============================================================
# SHEET 1: Functional Requirements
# ============================================================
ws1 = wb.active
ws1.title = "Functional Requirements"
ws1.sheet_properties.tabColor = "2D4285"

FR_DATA = [
    ["FR001", "Document Upload", "System shall support document upload via web interface from desktop and mobile devices", "High", "✅ Done", "Upload page at /upload supports file selection and upload. Responsive CSS works on mobile browsers. Accepts PNG, JPEG, PDF, TIFF, DOCX, HEIC, WebP. Max 16 MB."],
    ["FR002", "Document Upload", "System shall support batch upload of multiple documents simultaneously", "Medium", "✅ Done", "Upload page supports multiple file input. Files uploaded sequentially with per-file progress bars and overall progress."],
    ["FR003", "Document Upload", "System shall provide real-time image quality feedback during upload", "High", "❌ Not Done", "No DPI check or image quality scoring during upload. OCR confidence shown after extraction, but no pre-upload quality feedback."],
    ["FR004", "Document Classification", "System shall automatically classify document types (invoice, packing list, bill of lading, warehouse receipt)", "High", "✅ Done", "Score-based classifier in classify_document() identifies all 4 types using keyword scoring with confidence score."],
    ["FR005", "OCR Extraction", "System shall extract text data from uploaded document images using OCR technology", "High", "✅ Done", "Tesseract 5.5.2 with OEM 3, PSM 6 at 300 DPI. DOCX files parsed directly via python-docx (no OCR needed)."],
    ["FR006", "OCR Extraction", "System shall support multilingual OCR for Vietnamese and English text", "High", "✅ Done", "Default lang parameter is eng+vie. Tesseract trained data for both languages installed. Vietnamese diacritical marks handled."],
    ["FR007", "Field Recognition", "System shall identify and extract key fields (invoice number, date, amount, supplier name, line items, etc.)", "High", "✅ Done", "extract_fields() extracts 20+ field types: invoice_number, date, po_number, bl_number, total_amount, supplier_name, buyer_name, tax_code, hs_codes, vessel, ports, weights, containers, etc."],
    ["FR008", "Confidence Scoring", "System shall provide confidence scores (0-100%) for each extracted field", "High", "✅ Done", "Every field returned with confidence value (0-100). OCR page-level avg confidence also computed via image_to_data."],
    ["FR009", "Data Validation", "System shall validate extracted data against business rules (format, mandatory fields, data ranges)", "High", "✅ Done", "Automated validation runs after OCR: mandatory field checks per doc type (invoice needs invoice_number/date/total_amount/supplier_name, etc.), date format validation, amount numeric check, invoice number format check. Results stored in validation_results and displayed on document detail. Failed checks auto-flag the document."],
    ["FR010", "Cross-Verification", "System shall cross-check extracted data against related documents in the system", "High", "✅ Done", "Cross-document verification implemented. Users assign a Shipment Reference at upload time to group related documents (invoice, packing list, B/L, warehouse receipt) into a set. Cross-verify compares 11 overlapping fields (supplier, buyer, amounts, vessel, ports, weights, containers, packages, currency) only within the same shipment group. Smart comparison handles container subsets and unit text differences. Mismatches stored in validation_results (check_type='cross_verify'), auto-flag document, and trigger SNS notification with details."],
    ["FR011", "Reference Data Validation", "System shall validate extracted data against master data (supplier database, product catalog, HS codes)", "Medium", "❌ Not Done", "No master data tables. No supplier database, product catalog, or HS code reference validation."],
    ["FR012", "External API Integration", "System shall integrate with government customs APIs for real-time validation", "Medium", "❌ Not Done", "No external customs API integration. Would require Vietnam customs portal API access."],
    ["FR013", "Exception Handling", "System shall flag documents with low confidence scores (<70%) or validation failures for manual review", "High", "✅ Done", "Documents with any field confidence <70% or type_confidence <70% are automatically set to 'flagged' status. Validation failures also trigger auto-flagging. Flagged documents routed to review queue with reason logged in audit trail."],
    ["FR014", "Review Queue", "System shall provide a prioritized queue interface for manual review of flagged documents", "High", "✅ Done", "Review page shows documents with status flagged/reviewing/extracted. Displays doc type, confidence, issue count. Sorted by creation date."],
    ["FR015", "Review Interface", "System shall display original document image side-by-side with extracted data for reviewer validation", "High", "✅ Done", "Document detail page shows original document on the left (PDF in iframe, images inline via S3 presigned URL) and extracted fields on the right in a two-column layout. DOCX files show download link since browser preview is not supported."],
    ["FR016", "Manual Correction", "System shall allow authorized users to correct extraction errors and override validation flags", "High", "✅ Done", "Document detail has editable input fields. update_field endpoint saves corrected_value and marks reviewed=TRUE."],
    ["FR017", "Approval Workflow", "System shall support one-click approval for high-confidence extractions", "Medium", "✅ Done", "Approve & Send to SAP button. Single click triggers approval, updates status to completed, logs SAP push payload."],
    ["FR018", "ERP Integration", "System shall automatically push validated data to SAP ERP system via API", "High", "⚠️ Partial", "SAP integration simulated. Approved documents generate JSON payload in MIRO format. No actual SAP API connection — simulation/demo."],
    ["FR019", "WMS Integration", "System shall automatically push validated warehouse data to WMS via API", "High", "⚠️ Partial", "WMS integration simulated. Approved warehouse receipts/B/L in MIGO format. No actual WMS API connection — simulation only."],
    ["FR020", "Transaction Management", "System shall implement atomic transactions with rollback capability", "High", "⚠️ Partial", "PostgreSQL transactions used. No explicit rollback on partial failure — S3 upload not rolled back on DB failure. No distributed transaction management."],
    ["FR021", "Audit Trail", "System shall log all document processing activities", "High", "✅ Done", "audit_log table records actions with document_id, action, new_value, logged_at, user_id. Approval/rejection logged. Displayed in document detail."],
    ["FR022", "Error Analytics", "System shall track and categorize error types for continuous improvement", "Medium", "✅ Done", "Analytics page at /analytics shows: validation errors by check type (bar chart), errors by document type, manual corrections by doc type, recent validation failures table, and KPI cards (total docs, total errors, corrections, pass rate). Errors categorized as mandatory/date/amount/invoice types."],
    ["FR023", "User Dashboard", "System shall provide role-based dashboards showing relevant metrics", "Medium", "⚠️ Partial", "Dashboard shows 4 KPI cards (Total Documents, Processed Today, Avg Confidence, Flagged for Review) plus per-status counts and recent documents table. Login required with user name displayed. Not fully role-based — all roles see same dashboard view."],
    ["FR024", "Search Functionality", "System shall allow users to search processed documents by multiple criteria", "Medium", "✅ Done", "Dashboard has search bar (by doc ID/filename), status dropdown filter, document type dropdown filter, and Search/Clear buttons. Results returned instantly with up to 100 documents."],
    ["FR025", "Document Retrieval", "System shall allow users to retrieve and view original documents", "Medium", "✅ Done", "Document detail page shows original document preview (PDF/image via S3 presigned URL) and a 'Download Original' button that streams the file from S3 to the browser."],
    ["FR026", "Notification System", "System shall send notifications for documents requiring review", "Low", "✅ Done", "AWS SNS email notifications sent to david@g-asiapac.com.vn when documents are flagged for review, approved, or rejected. Includes document ID, reason, and link to review page."],
    ["FR027", "User Management", "System shall support role-based access control with different permission levels", "High", "✅ Done", "Login page with session-based authentication. 4 users with 3 roles: admin (full access), reviewer (review/approve/reject/analytics/reports/export), uploader (dashboard + upload only). Role enforced at route level via @role_required decorator. Nav bar hides inaccessible pages per role. Approve/reject/edit buttons hidden for uploaders. 403 page shown for unauthorized access attempts. Passwords hashed with SHA-256."],
    ["FR028", "ML Model Training", "System shall capture human corrections for model retraining", "Medium", "❌ Not Done", "Corrections saved (corrected_value) but no ML retraining pipeline exists."],
    ["FR029", "Reporting", "System shall generate standard reports on processing volume, accuracy, errors, cost savings", "Medium", "✅ Done", "Reports page at /report with 4 downloadable reports: (1) Full Processing Report PDF — volume by type/status, accuracy metrics, error analysis, cost savings estimate, daily volume table; (2) Document Summary CSV; (3) Full Data Export CSV; (4) Full Data Export JSON. PDF generated via fpdf2."],
    ["FR030", "Data Export", "System shall allow export of extracted data in CSV, Excel, JSON", "Low", "✅ Done", "Dashboard has Export CSV and Export JSON buttons. CSV export includes all documents with extracted fields. JSON export includes structured document data with field values and confidence scores. Files download directly."],
]

NFR_DATA = [
    ["NFR001", "Performance", "OCR extraction within 10 seconds per document", "≤10s per doc", "⚠️ Partial", "DOCX <2s. PDF OCR 5-30s depending on pages. Single-page PDFs likely within 10s; multi-page may exceed."],
    ["NFR002", "Performance", "Support 500 concurrent users", "500+ users", "❌ Not Done", "Flask dev server, single-threaded. No WSGI server (gunicorn). Would not handle 500 concurrent users."],
    ["NFR003", "Performance", "UI response time under 2 seconds", "≤2s", "✅ Done", "Static HTML templates with minimal JS. Simple SQL queries. Fast page loads for normal usage."],
    ["NFR004", "Performance", "Handle 15,000 documents per day", "15,000/day", "❌ Not Done", "Single EC2 for OCR, single for web. No queue system. Sequential processing. Cannot sustain 15K docs/day."],
    ["NFR005", "Scalability", "Horizontal scaling for 3x volume growth", "3x capacity", "❌ Not Done", "Single EC2 instances, no ASG, no ALB. Not horizontally scalable."],
    ["NFR006", "Scalability", "Auto-scale based on queue depth", "Auto-scaling", "❌ Not Done", "No ASG, no SQS queue. Fixed EC2 instances."],
    ["NFR007", "Availability", "99.5% uptime during business hours", "≥99.5%", "⚠️ Partial", "Systemd Restart=on-failure for process recovery. RDS managed. But single EC2 with no redundancy."],
    ["NFR008", "Availability", "Automated failover for critical components", "<5 min failover", "❌ Not Done", "No failover. Single AZ. No standby, no Route53 health checks, no ALB."],
    ["NFR009", "Reliability", "Zero data loss during processing failures", "0% data loss", "⚠️ Partial", "S3 upload before OCR preserves raw docs. PostgreSQL ACID. But no explicit error recovery for mid-process crashes."],
    ["NFR010", "Reliability", "Transaction rollback for failed insertions", "100% rollback", "⚠️ Partial", "PostgreSQL transactions with commit. No explicit try/except/rollback in all paths. S3 not rolled back on DB failure."],
    ["NFR011", "Security", "Encrypt data in transit using TLS 1.3+", "TLS 1.3+", "✅ Done", "HTTPS via Let's Encrypt (certbot) on Nginx. SSL active for idp.pngha.io.vn."],
    ["NFR012", "Security", "Encrypt data at rest using AES-256", "AES-256", "⚠️ Partial", "RDS has AWS default encryption. S3 uses SSE-S3. EC2 EBS may not have explicit encryption."],
    ["NFR013", "Security", "Multi-factor authentication for all users", "100% MFA", "⚠️ Partial", "Login page with username/password authentication implemented. Session-based auth with SHA-256 password hashing. MFA not yet implemented."],
    ["NFR014", "Security", "Role-based access control (RBAC)", "RBAC enforced", "✅ Done", "3 roles enforced: admin (full access including SAP), reviewer (review/approve/reject/analytics/reports/export), uploader (dashboard + upload only). @role_required decorator on all restricted routes. Nav bar and action buttons adapt per role. 403 page for unauthorized access."],
    ["NFR015", "Security", "Comprehensive audit logs for all actions", "100% logging", "✅ Done", "audit_log captures all actions: upload, field corrections (with old→new values), auto_flagged, validation_failed, approved_sap_push, rejected. All entries include username, timestamp, document_id, and detail."],
    ["NFR016", "Security", "Automated security scanning", "Weekly scans", "❌ Not Done", "No security scanning tools. No vulnerability detection."],
    ["NFR017", "Compliance", "Audit trails for 7+ years", "7 year retention", "⚠️ Partial", "Logs in PostgreSQL with no auto-deletion. No formal 7-year retention policy."],
    ["NFR018", "Compliance", "Immutable audit logs", "Immutable logging", "❌ Not Done", "Regular PostgreSQL table — can be updated/deleted. Not write-once/append-only."],
    ["NFR019", "Compliance", "Vietnam customs and tax documentation support", "Full compliance", "⚠️ Partial", "Extracts tax code, HS codes, customs fields. Vietnamese text supported. No formal compliance review."],
    ["NFR020", "Usability", "Intuitive UI, minimal training (<2 hours)", "≤2h training", "✅ Done", "Clean simple UI. Upload → Review → Approve workflow is straightforward."],
    ["NFR021", "Usability", "Mobile-responsive interface", "Full mobile", "⚠️ Partial", "Flexible CSS layouts. Basic responsiveness. No explicit @media queries; tables may overflow on small screens."],
    ["NFR022", "Usability", "Vietnamese and English language UI", "2 languages", "❌ Not Done", "UI is English-only. No language switcher. OCR supports Vietnamese but UI labels are English."],
    ["NFR023", "Usability", "Context-sensitive help and tooltips", "Help available", "❌ Not Done", "No tooltips, no help text, no user guide in the application."],
    ["NFR024", "Maintainability", "Modular architecture, zero-downtime updates", "Zero-downtime", "⚠️ Partial", "Separate EC2 for OCR/Web allows independent updates. But systemctl restart causes brief downtime."],
    ["NFR025", "Maintainability", "Test coverage ≥80%", "≥80% coverage", "❌ Not Done", "No test suite. Zero test coverage."],
    ["NFR026", "Maintainability", "Comprehensive technical documentation", "Complete docs", "⚠️ Partial", "README, DEPLOYMENT_SUMMARY, QUICK_START exist. No API docs or code-level documentation."],
    ["NFR027", "Interoperability", "RESTful APIs with JSON", "REST API", "✅ Done", "Flask REST endpoints for upload, approve, reject, update_field, status, OCR extract, health. All return JSON."],
    ["NFR028", "Interoperability", "SAP ERP integration via SAP APIs/RFC", "SAP integrated", "⚠️ Partial", "SAP simulated — JSON payloads in MIRO format. No actual SAP RFC/BAPI connection."],
    ["NFR029", "Interoperability", "WMS integration via standard APIs", "WMS integrated", "⚠️ Partial", "WMS simulated — MIGO format display. No actual WMS API connection."],
    ["NFR030", "Disaster Recovery", "Automated daily backups, RPO ≤24h", "RPO ≤24h", "⚠️ Partial", "RDS automated backups (7-day). S3 durable. EC2 has no automated backup/AMI snapshots."],
    ["NFR031", "Disaster Recovery", "RTO of 4 hours", "RTO ≤4h", "⚠️ Partial", "RDS restorable from snapshot. EC2 recreatable from deploy scripts. No documented/tested DR procedure."],
    ["NFR032", "Monitoring", "Real-time monitoring dashboard", "Real-time", "⚠️ Partial", "/api/status endpoint returns DB status and doc stats. No CloudWatch/Grafana dashboard."],
    ["NFR033", "Monitoring", "Automated alerts for anomalies", "Auto alerting", "⚠️ Partial", "AWS SNS email notifications for flagged documents, approvals, and rejections sent to subscribed email. No CloudWatch alarms for system-level metrics (CPU, memory, disk) yet."],
    ["NFR034", "Accuracy", "OCR ≥85% character-level accuracy", "≥85%", "✅ Done", "Tesseract 5.5.2 LSTM at 300 DPI. DOCX gets 99%. PDF OCR meets 85%+ for clean documents."],
    ["NFR035", "Accuracy", "Field extraction ≥90% accuracy", "≥90%", "⚠️ Partial", "Key fields extract well from DOCX. PDF OCR has multi-column challenges. 20+ regex patterns. Accuracy varies."],
    ["NFR036", "Accuracy", "Detect ≥95% data mismatches", "≥95% detection", "⚠️ Partial", "Cross-document verification detects mismatches in 11 overlapping fields within user-defined shipment groups. Smart comparison handles container subsets and unit text. Detection rate depends on OCR extraction quality — clean DOCX achieves near 100%, scanned PDFs may miss fields."],
    ["NFR037", "Capacity", "150GB annual data growth", "150GB+/year", "✅ Done", "S3 for documents (unlimited). RDS scalable. Supports capacity requirement."],
    ["NFR038", "Capacity", "2.5M+ document records annually", "2.5M+/year", "⚠️ Partial", "PostgreSQL can handle millions of records. No indexing optimization or partitioning verified."],
    ["NFR039", "Localization", "Handle Vietnamese diacritical marks", "Full Vietnamese", "✅ Done", "Tesseract vie pack. PostgreSQL UTF-8. Python 3 Unicode. Vietnamese characters preserved."],
    ["NFR040", "Cost Efficiency", "Infrastructure costs <$0.01/document", "≤$0.01/doc", "⚠️ Partial", "~$150-200/month. At 10K docs/month = $0.015-0.02/doc. Meets target at higher volumes only."],
]

# --- Build Functional Requirements sheet ---
fr_headers = ["Req ID", "Category", "Requirement Description", "Priority", "Status", "Explanation"]
fr_widths = [10, 22, 55, 10, 14, 65]

ws1.merge_cells("A1:F1")
ws1["A1"] = "Section 3.1 — Functional Requirements Checklist"
ws1["A1"].font = TITLE_FONT
ws1["A1"].alignment = Alignment(vertical="center")
ws1.row_dimensions[1].height = 30

for col_idx, (header, width) in enumerate(zip(fr_headers, fr_widths), 1):
    cell = ws1.cell(row=2, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

for row_idx, row_data in enumerate(FR_DATA, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        if col_idx == 1:
            cell.font = BOLD_FONT
            cell.alignment = CENTER
        elif col_idx == 4:
            cell.alignment = CENTER
        elif col_idx == 5:
            fill, font = status_style(value)
            cell.fill = fill
            cell.font = font
            cell.alignment = CENTER

# --- Build Non-Functional Requirements sheet ---
ws2 = wb.create_sheet("Non-Functional Requirements")
ws2.sheet_properties.tabColor = "2D4285"

nfr_headers = ["Req ID", "Category", "Requirement Description", "Target Metric", "Status", "Explanation"]
nfr_widths = [10, 18, 55, 18, 14, 65]

ws2.merge_cells("A1:F1")
ws2["A1"] = "Section 3.2 — Non-Functional Requirements Checklist"
ws2["A1"].font = TITLE_FONT
ws2["A1"].alignment = Alignment(vertical="center")
ws2.row_dimensions[1].height = 30

for col_idx, (header, width) in enumerate(zip(nfr_headers, nfr_widths), 1):
    cell = ws2.cell(row=2, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

for row_idx, row_data in enumerate(NFR_DATA, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        if col_idx == 1:
            cell.font = BOLD_FONT
            cell.alignment = CENTER
        elif col_idx == 4:
            cell.alignment = CENTER
        elif col_idx == 5:
            fill, font = status_style(value)
            cell.fill = fill
            cell.font = font
            cell.alignment = CENTER

# --- Build Summary sheet ---
ws3 = wb.create_sheet("Summary")
ws3.sheet_properties.tabColor = "4CAF50"

ws3.merge_cells("A1:E1")
ws3["A1"] = "Requirements Checklist — Overall Summary"
ws3["A1"].font = TITLE_FONT
ws3.row_dimensions[1].height = 30

sum_headers = ["Category", "Total", "✅ Done", "⚠️ Partial", "❌ Not Done"]
sum_widths = [35, 10, 14, 14, 14]
for col_idx, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
    cell = ws3.cell(row=3, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws3.column_dimensions[get_column_letter(col_idx)].width = w

# Count statuses
fr_done = sum(1 for r in FR_DATA if "Done" in r[4] and "Not" not in r[4])
fr_partial = sum(1 for r in FR_DATA if "Partial" in r[4])
fr_not = sum(1 for r in FR_DATA if "Not Done" in r[4])
nfr_done = sum(1 for r in NFR_DATA if "Done" in r[4] and "Not" not in r[4])
nfr_partial = sum(1 for r in NFR_DATA if "Partial" in r[4])
nfr_not = sum(1 for r in NFR_DATA if "Not Done" in r[4])

summary_rows = [
    ["Functional Requirements (FR001-FR030)", 30, fr_done, fr_partial, fr_not],
    ["Non-Functional Requirements (NFR001-NFR040)", 40, nfr_done, nfr_partial, nfr_not],
    ["TOTAL", 70, fr_done+nfr_done, fr_partial+nfr_partial, fr_not+nfr_not],
]

for row_idx, row_data in enumerate(summary_rows, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if row_idx == 6:
            cell.font = BOLD_FONT
        else:
            cell.font = NORMAL_FONT
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        if col_idx == 3:
            cell.fill = DONE_FILL
        elif col_idx == 4:
            cell.fill = PARTIAL_FILL
        elif col_idx == 5:
            cell.fill = NOT_DONE_FILL

# Percentage row
ws3.cell(row=8, column=1, value="Completion Percentages").font = BOLD_FONT
for row_idx, row_data in enumerate(summary_rows, 9):
    total = row_data[1]
    ws3.cell(row=row_idx, column=1, value=row_data[0]).font = NORMAL_FONT
    ws3.cell(row=row_idx, column=1).border = THIN_BORDER
    for col_idx, val in enumerate(row_data[2:], 3):
        pct = f"{val/total*100:.0f}%"
        cell = ws3.cell(row=row_idx, column=col_idx, value=pct)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if col_idx == 3: cell.fill = DONE_FILL
        elif col_idx == 4: cell.fill = PARTIAL_FILL
        elif col_idx == 5: cell.fill = NOT_DONE_FILL

# Key gaps
ws3.cell(row=13, column=1, value="Key Gaps (High Priority Not Done)").font = Font(name="Arial", size=12, bold=True, color="D32F2F")
gaps = [
    "FR003 — No image quality feedback during upload",
    "NFR002 — Cannot handle 500 concurrent users (single Flask dev server)",
    "NFR005/006 — No horizontal scaling or auto-scaling",
    "NFR008 — No automated failover",
    "NFR013 — MFA not yet implemented (login exists but no MFA)",
    "NFR018 — Audit logs not immutable",
]
for i, gap in enumerate(gaps, 14):
    cell = ws3.cell(row=i, column=1, value=gap)
    cell.font = Font(name="Arial", size=11, color="D32F2F")

# Freeze panes
ws1.freeze_panes = "A3"
ws2.freeze_panes = "A3"

# Auto-filter
ws1.auto_filter.ref = f"A2:F{len(FR_DATA)+2}"
ws2.auto_filter.ref = f"A2:F{len(NFR_DATA)+2}"

wb.save("REQUIREMENTS_CHECKLIST.xlsx")
print("Created REQUIREMENTS_CHECKLIST.xlsx")
